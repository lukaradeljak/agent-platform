"""
Stage 4a: Excel Generation.
Creates a professionally formatted Excel file with enriched lead data.
"""

import json
import logging
from datetime import date

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tools.config import TMP_DIR

logger = logging.getLogger("pipeline")


def _parse_automations(lead):
    """Parse automation_suggestions JSON into individual columns."""
    raw = lead.get("automation_suggestions", "")
    if not raw:
        return ["", "", ""]

    try:
        if isinstance(raw, str):
            automations = json.loads(raw)
        else:
            automations = raw
    except (json.JSONDecodeError, TypeError):
        return ["", "", ""]

    results = []
    for auto in automations[:3]:
        if isinstance(auto, dict):
            name = auto.get("name", "")
            desc = auto.get("description", "")
            value = auto.get("value", "")
            results.append(f"{name}: {desc} ({value})")
        else:
            results.append(str(auto))

    # Pad to 3 entries
    while len(results) < 3:
        results.append("")

    return results[:3]


def run(leads, run_date=None):
    """
    Generate an Excel file from the enriched leads.
    Returns the file path of the generated Excel.
    """
    if run_date is None:
        run_date = date.today().isoformat()

    if not leads:
        logger.warning("No leads to write to Excel.")
        return None

    logger.info(f"Building Excel with {len(leads)} leads...")

    # Build data for DataFrame
    rows = []
    for lead in leads:
        autos = _parse_automations(lead)
        rows.append({
            "Empresa": lead.get("company_name", ""),
            "Contacto": lead.get("contact_name", ""),
            "Email": lead.get("email", ""),
            "Website": lead.get("website", ""),
            "Telefono": lead.get("phone", ""),
            "Ciudad": lead.get("city", ""),
            "Pais": lead.get("country", ""),
            "Resumen": lead.get("ai_summary", ""),
            "Automatizacion 1": autos[0],
            "Automatizacion 2": autos[1],
            "Automatizacion 3": autos[2],
            "Fecha": lead.get("discovered_date", run_date),
        })

    df = pd.DataFrame(rows)

    # Save to Excel
    TMP_DIR.mkdir(exist_ok=True)
    file_path = TMP_DIR / f"leads_{run_date}.xlsx"

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")
        ws = writer.sheets["Leads"]
        _apply_formatting(ws, len(rows))

    logger.info(f"Excel saved: {file_path}")
    return str(file_path)


def _apply_formatting(ws, row_count):
    """Apply professional formatting to the worksheet."""
    # Colors
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    alt_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    normal_font = Font(name="Calibri", size=10)
    border = Border(
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Format header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format data rows
    for row in range(2, row_count + 2):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Alternate row shading
            if row % 2 == 0:
                cell.fill = alt_fill

    # Column widths (approximate)
    column_widths = {
        "A": 25,  # Empresa
        "B": 20,  # Contacto
        "C": 28,  # Email
        "D": 30,  # Website
        "E": 18,  # Telefono
        "F": 15,  # Ciudad
        "G": 15,  # Pais
        "H": 45,  # Resumen
        "I": 50,  # Auto 1
        "J": 50,  # Auto 2
        "K": 50,  # Auto 3
        "L": 12,  # Fecha
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Make Website column clickable (hyperlinks)
    for row in range(2, row_count + 2):
        cell = ws.cell(row=row, column=4)  # Column D = Website
        url = cell.value
        if url and isinstance(url, str) and url.startswith("http"):
            cell.hyperlink = url
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")


if __name__ == "__main__":
    # Test with sample data
    sample_leads = [
        {
            "company_name": "Agencia Test",
            "contact_name": "Juan Perez",
            "email": "juan@test.com",
            "website": "https://test.com",
            "phone": "+34 91 123 4567",
            "city": "Madrid",
            "country": "Espana",
            "ai_summary": "Agencia de marketing digital especializada en SEO y SEM.",
            "automation_suggestions": json.dumps([
                {"name": "Auto 1", "description": "Desc 1", "value": "Val 1"},
                {"name": "Auto 2", "description": "Desc 2", "value": "Val 2"},
                {"name": "Auto 3", "description": "Desc 3", "value": "Val 3"},
            ]),
            "discovered_date": "2026-02-02",
        }
    ]
    path = run(sample_leads)
    print(f"Test Excel: {path}")
